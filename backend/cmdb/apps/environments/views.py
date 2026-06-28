"""Views for environments app."""
import csv
import json
from io import BytesIO
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db import connection
from django.db.models import Q, F, FloatField, Count
from django.db.models.functions import Cast, Coalesce
from django_tables2 import SingleTableView, RequestConfig
from django_filters.views import FilterView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Environment, EnvironmentDependency
from .tables import (
    EnvironmentTable, TeamAggregationTable, CloudRegionTable,
    ControllerTable, DependencyHotspotTable, VersionComplianceTable,
    ServicePrimitiveTable
)
from .filters import EnvironmentFilter


class EnvironmentListView(FilterView, SingleTableView):
    """
    Main CMDB list view: filterable, sortable, paginated table of environments.
    """
    model = Environment
    table_class = EnvironmentTable
    filterset_class = EnvironmentFilter
    template_name = 'environments/list.html'
    paginate_by = 50

    # Sentinel sort key so envs without live placement always sort to the
    # bottom (ascending) regardless of their name.
    _NO_PLACEMENT_SORT = '￿'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        # Keep the landing page fast: with ~2k environments we don't load the
        # full table on a bare visit. The table stays empty until the user
        # interacts — submits the filter form or picks a region tab — i.e. as
        # soon as there is any query parameter on the request.
        self.show_results = bool(request.GET)
        # Populated only when sorting by placement (see get_table_data); lets
        # get_context_data reuse the already-fetched placement for the page.
        self._placement_cache = None

    # Columns whose values come from Redis placement, not the DB. Sorting by
    # any of them requires materialising the set and ordering in memory.
    _REDIS_SORT_COLUMNS = {'placement_status', 'vms', 'ha', 'resilient'}

    def _resilient_names(self):
        """Resilient env names (Redis-derived), cached per request."""
        if not hasattr(self, '_resilient_cache'):
            from cmdb.redis_client import resilient_env_names
            self._resilient_cache = resilient_env_names()
        return self._resilient_cache

    def _sorting_by_redis_column(self) -> bool:
        """True when the active ordering is a Redis-derived column."""
        sort = self.request.GET.get('sort', '')
        fields = [s.strip().lstrip('-') for s in sort.split(',') if s.strip()]
        return any(f in self._REDIS_SORT_COLUMNS for f in fields)

    def get_queryset(self):
        """Get queryset filtered by region tab if present.

        Returns an empty queryset until the user has applied a filter, so the
        landing page renders instantly without scanning every environment.
        """
        qs = super().get_queryset()

        if not self.show_results:
            return qs.none()

        # Tab filter: 'modernize' = not-resilient envs; otherwise region.
        tab = self.request.GET.get('tab')
        if tab == 'modernize':
            qs = qs.exclude(name__in=self._resilient_names())
        elif tab and tab != 'all':
            qs = qs.filter(region=tab)

        # Service deep-link (?service=dbaas|ck8s|jenkins|builders): restrict to
        # that service's envs so service pages can link "DBaaS for <team>".
        service = self.request.GET.get('service')
        if service:
            sq = service_base_queryset(service)
            if sq is not None:
                qs = qs.filter(pk__in=sq.values('pk'))

        # ?no_team=1: envs with no team (the teams page's synthetic 'Unknown').
        if self.request.GET.get('no_team'):
            qs = qs.filter(Q(team__isnull=True) | Q(team=''))

        # The Quota columns sort by their numeric value. Annotate a numeric
        # cast (per column actually sorted) so ordering is numeric rather than
        # lexicographic over the JSON value. Prefixed names avoid clashing with
        # Environment.quota_* model fields.
        sort = self.request.GET.get('sort', '')
        quota_ann = {}
        if 'quota_vcpu' in sort:
            quota_ann['sort_cores'] = Cast(F('quotas__cores'), FloatField())
        if 'quota_ram' in sort:
            quota_ann['sort_ram'] = Cast(F('quotas__ram'), FloatField())
        if 'quota_disk' in sort:
            quota_ann['sort_disk'] = Cast(F('quotas__gigabytes'), FloatField())
        if quota_ann:
            qs = qs.annotate(**quota_ann)

        # "Consumed by" shows consumer_team or team; order by the same fallback.
        if 'consumed_by' in sort:
            qs = qs.annotate(consumed_by_sort=Coalesce('requester', 'team'))

        # Default row order when no column sort is active: most recently
        # updated. django-tables2 overrides this when a column header is used.
        qs = qs.order_by('-updated_at')

        return qs

    def get_table_data(self):
        """Table data for the current request.

        Placement, VM count and resilience all come from Redis, not the DB, so
        they can't be ordered with the queryset. When the user sorts by one of
        those columns we materialise the filtered set, fetch placement for every
        row in one pipelined call, and stamp sort keys on each instance.
        Returning a list makes django-tables2 sort in memory (TableListData) by
        those keys. Any other ordering keeps the lazy, paginated queryset path.
        """
        data = super().get_table_data()
        if not (self.show_results and self._sorting_by_redis_column()):
            return data

        envs = list(data)
        self._placement_cache = self._placement_map([e.name for e in envs])
        for env in envs:
            placement = self._placement_cache.get(env.name) or {}
            host = placement.get('primary_host')
            vm_count = placement.get('vm_count') or 0
            node_count = len(placement.get('hosts') or [])
            env.placement_sort = host.lower() if host else self._NO_PLACEMENT_SORT
            env.vms_sort = vm_count
            env.ha_sort = int(vm_count > 2)
            env.resilient_sort = int(
                bool(env.gitops_managed) and vm_count > 3 and node_count > 1
            )
        return envs

    @staticmethod
    def _placement_map(names):
        """Return ``{env_name: placement_dict}`` for the given names.

        Fetches only the visible rows in a single pipelined round trip — never
        the whole table — so this stays cheap regardless of result-set size.
        Keyed by name (not instance) because django-tables2 re-evaluates the
        page queryset on render, yielding fresh model instances each time.
        Mirrors ``redis_client.get_placement``'s ``ps5-`` prefix fallback.
        """
        from cmdb.redis_client import get_redis_client

        names = list(names)
        if not names:
            return {}

        client = get_redis_client()
        pipe = client.pipeline()
        for name in names:
            pipe.get(f"env:{name}:placement")
        # ps5 fallback: parser may have prefixed model names with 'ps5-'.
        fallback_pos = {}
        for i, name in enumerate(names):
            if not name.startswith('ps5-'):
                fallback_pos[i] = len(names) + len(fallback_pos)
                pipe.get(f"env:ps5-{name}:placement")
        results = pipe.execute()

        placement = {}
        for i, name in enumerate(names):
            raw = results[i]
            if raw is None and i in fallback_pos:
                raw = results[fallback_pos[i]]
            try:
                placement[name] = json.loads(raw) if raw else None
            except (json.JSONDecodeError, TypeError):
                placement[name] = None
        return placement

    def get_table_kwargs(self):
        """Pass the under-maintenance env ids to the table for the badge (#38)."""
        kwargs = super().get_table_kwargs()
        from cmdb.apps.maintenance.queries import environment_ids_under_maintenance
        kwargs['under_maintenance_ids'] = environment_ids_under_maintenance()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Whether to render the results table (see setup()).
        context['show_results'] = self.show_results
        context['total_count'] = Environment.objects.count()

        # Live placement for just the rows on the current page, keyed by env
        # name. The table looks each row up by name in render_placement_status.
        if self.show_results:
            table = context.get('table')
            page = getattr(table, 'page', None)
            if page is not None and table is not None:
                names = [row.record.name for row in page.object_list]
                if self._placement_cache is not None:
                    # Already fetched for the whole set to sort by placement.
                    table.placement_by_name = {
                        n: self._placement_cache.get(n) for n in names
                    }
                else:
                    table.placement_by_name = self._placement_map(names)

        # Add region tabs + the Modernize (not-resilient) tab.
        modernize_count = context['total_count'] - len(self._resilient_names())
        context['tabs'] = [
            {'name': 'All', 'value': 'all', 'count': context['total_count']},
            {'name': 'AMER', 'value': 'amer', 'count': Environment.objects.filter(region='amer').count()},
            {'name': 'EMEA', 'value': 'emea', 'count': Environment.objects.filter(region='emea').count()},
            {'name': 'APAC', 'value': 'apac', 'count': Environment.objects.filter(region='apac').count()},
            {'name': 'Modernize', 'value': 'modernize', 'count': modernize_count},
        ]
        context['active_tab'] = self.request.GET.get('tab', 'all')

        # Add user email for "My Environments" shortcut
        # In production, this could come from SSO/authentication
        # For now, users can bookmark their custom filter URL
        context['user_email'] = 'your.email'

        return context


def environment_detail(request, name):
    """Detail view for a single environment."""
    from cmdb.redis_client import get_placement, cloud_has_placement

    env = get_object_or_404(Environment, name=name)

    # Get dependencies (what this env depends on)
    dependencies = EnvironmentDependency.objects.filter(
        environment_name=name
    ).select_related()

    # Get dependents (what depends on this env)
    dependents = EnvironmentDependency.objects.filter(
        depends_on_name=name
    ).select_related()

    # Get live placement from Redis
    placement = get_placement(env.name)
    cloud_supports_placement = cloud_has_placement(env.cloud) if env.cloud else False

    # Active/upcoming maintenance on this env's nodes (#38).
    from cmdb.apps.maintenance.queries import windows_for_node
    maintenance_window = None
    for node in (env.primary_node, env.secondary_node):
        if node:
            windows = windows_for_node(node)
            if windows:
                maintenance_window = windows[0]
                break

    # Storage resources this environment accesses (#47).
    from cmdb.apps.storage.models import StorageEnvironmentAccess
    storage_accesses = (
        StorageEnvironmentAccess.objects.filter(environment=env)
        .select_related('storage')
        .order_by('storage__name')
    )

    # CIA ownership fallback: when an env has no CIA data, surface who edited its
    # definition file in is-infrastructure so the owner can be tracked down (#cia-fallback).
    env_has_cia = bool(env.cia_owner or env.cia_risk_owner or env.cia_custodian)
    cia_editors = None
    if not env_has_cia and env.git_path:
        from cmdb.integrations import infra_history
        cia_editors = infra_history.file_editors(env.git_path)

    context = {
        'environment': env,
        'dependencies': dependencies,
        'dependents': dependents,
        'placement': placement,
        'cloud_supports_placement': cloud_supports_placement,
        'maintenance_window': maintenance_window,
        'storage_accesses': storage_accesses,
        'env_has_cia': env_has_cia,
        'cia_editors': cia_editors,
    }

    return render(request, 'environments/detail.html', context)


def export_csv(request):
    """Export environments to CSV, respecting list view filters."""
    # Apply filters from EnvironmentFilter
    filterset = EnvironmentFilter(request.GET, queryset=Environment.objects.all())
    queryset = filterset.qs

    # Apply region tab filter if present
    region_filter = request.GET.get('tab')
    if region_filter and region_filter != 'all':
        queryset = queryset.filter(region=region_filter)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="environments.csv"'

    writer = csv.writer(response)

    # Write header row with all fields
    header = [
        'name', 'region', 'cloud', 'owner', 'team', 'env_type', 'status',
        'criticality_tier', 'data_classification', 'service_primitive', 'service_class',
        'juju_controller', 'juju_series', 'juju_controller_stage',
        'bastion_server', 'risk_group', 'cia_owner', 'cia_risk_owner', 'cia_custodian',
        'slo_level', 'slo_rto', 'live', 'description',
        'database_size', 'control_plane_size', 'control_plane_units',
        'network_size', 'compute_architecture', 'postgresql_major_version',
        'git_path', 'last_git_commit', 'created_at', 'end_date', 'updated_at'
    ]
    writer.writerow(header)

    # Write data rows
    for env in queryset:
        writer.writerow([
            env.name,
            env.region or '',
            env.cloud or '',
            env.owner or '',
            env.team or '',
            env.env_type or '',
            env.status or '',
            env.criticality_tier if env.criticality_tier is not None else '',
            env.data_classification or '',
            env.service_primitive or '',
            env.service_class or '',
            env.juju_controller or '',
            env.juju_series or '',
            env.juju_controller_stage or '',
            env.bastion_server or '',
            env.risk_group or '',
            env.cia_owner or '',
            env.cia_risk_owner or '',
            env.cia_custodian or '',
            env.slo_level or '',
            env.slo_rto if env.slo_rto is not None else '',
            env.live if env.live is not None else '',
            env.description or '',
            env.database_size or '',
            env.control_plane_size or '',
            env.control_plane_units if env.control_plane_units is not None else '',
            env.network_size if env.network_size is not None else '',
            env.compute_architecture or '',
            env.postgresql_major_version or '',
            env.git_path or '',
            env.last_git_commit or '',
            env.created_at.isoformat() if env.created_at else '',
            env.end_date.isoformat() if env.end_date else '',
            env.updated_at.isoformat() if env.updated_at else '',
        ])

    return response


def export_xls(request):
    """Export environments to Excel (XLS) with three sheets: Environments, Dependencies, Placement."""
    # Apply filters from EnvironmentFilter
    filterset = EnvironmentFilter(request.GET, queryset=Environment.objects.all())
    queryset = filterset.qs

    # Apply region tab filter if present
    region_filter = request.GET.get('tab')
    if region_filter and region_filter != 'all':
        queryset = queryset.filter(region=region_filter)

    # Create workbook
    wb = Workbook()

    # Sheet 1: Environments
    ws_envs = wb.active
    ws_envs.title = "Environments"

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write environment headers
    env_headers = [
        'name', 'region', 'cloud', 'owner', 'team', 'env_type', 'status',
        'criticality_tier', 'data_classification', 'service_primitive', 'service_class',
        'juju_controller', 'juju_series', 'juju_controller_stage',
        'bastion_server', 'risk_group', 'cia_owner', 'cia_risk_owner', 'cia_custodian',
        'slo_level', 'slo_rto', 'live', 'description',
        'database_size', 'control_plane_size', 'control_plane_units',
        'network_size', 'compute_architecture', 'postgresql_major_version',
        'git_path', 'last_git_commit', 'created_at', 'end_date', 'updated_at'
    ]

    for col_num, header in enumerate(env_headers, 1):
        cell = ws_envs.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write environment data
    for row_num, env in enumerate(queryset, 2):
        ws_envs.cell(row=row_num, column=1, value=env.name)
        ws_envs.cell(row=row_num, column=2, value=env.region or '')
        ws_envs.cell(row=row_num, column=3, value=env.cloud or '')
        ws_envs.cell(row=row_num, column=4, value=env.owner or '')
        ws_envs.cell(row=row_num, column=5, value=env.team or '')
        ws_envs.cell(row=row_num, column=6, value=env.env_type or '')
        ws_envs.cell(row=row_num, column=7, value=env.status or '')
        ws_envs.cell(row=row_num, column=8, value=env.criticality_tier if env.criticality_tier is not None else '')
        ws_envs.cell(row=row_num, column=9, value=env.data_classification or '')
        ws_envs.cell(row=row_num, column=10, value=env.service_primitive or '')
        ws_envs.cell(row=row_num, column=11, value=env.service_class or '')
        ws_envs.cell(row=row_num, column=12, value=env.juju_controller or '')
        ws_envs.cell(row=row_num, column=13, value=env.juju_series or '')
        ws_envs.cell(row=row_num, column=14, value=env.juju_controller_stage or '')
        ws_envs.cell(row=row_num, column=15, value=env.bastion_server or '')
        ws_envs.cell(row=row_num, column=16, value=env.risk_group or '')
        ws_envs.cell(row=row_num, column=17, value=env.cia_owner or '')
        ws_envs.cell(row=row_num, column=18, value=env.cia_risk_owner or '')
        ws_envs.cell(row=row_num, column=19, value=env.cia_custodian or '')
        ws_envs.cell(row=row_num, column=20, value=env.slo_level or '')
        ws_envs.cell(row=row_num, column=21, value=env.slo_rto if env.slo_rto is not None else '')
        ws_envs.cell(row=row_num, column=22, value=env.live if env.live is not None else '')
        ws_envs.cell(row=row_num, column=23, value=env.description or '')
        ws_envs.cell(row=row_num, column=24, value=env.database_size or '')
        ws_envs.cell(row=row_num, column=25, value=env.control_plane_size or '')
        ws_envs.cell(row=row_num, column=26, value=env.control_plane_units if env.control_plane_units is not None else '')
        ws_envs.cell(row=row_num, column=27, value=env.network_size if env.network_size is not None else '')
        ws_envs.cell(row=row_num, column=28, value=env.compute_architecture or '')
        ws_envs.cell(row=row_num, column=29, value=env.postgresql_major_version or '')
        ws_envs.cell(row=row_num, column=30, value=env.git_path or '')
        ws_envs.cell(row=row_num, column=31, value=env.last_git_commit or '')
        ws_envs.cell(row=row_num, column=32, value=env.created_at.isoformat() if env.created_at else '')
        ws_envs.cell(row=row_num, column=33, value=env.end_date.isoformat() if env.end_date else '')
        ws_envs.cell(row=row_num, column=34, value=env.updated_at.isoformat() if env.updated_at else '')

    # Auto-size columns
    for col_num in range(1, len(env_headers) + 1):
        ws_envs.column_dimensions[get_column_letter(col_num)].width = 15

    # Sheet 2: Dependencies
    ws_deps = wb.create_sheet(title="Dependencies")

    dep_headers = ['environment_name', 'depends_on_name', 'dependency_type']
    for col_num, header in enumerate(dep_headers, 1):
        cell = ws_deps.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Get dependencies for environments in queryset
    env_names = list(queryset.values_list('name', flat=True))
    dependencies = EnvironmentDependency.objects.filter(
        environment_name__in=env_names
    ).select_related()

    for row_num, dep in enumerate(dependencies, 2):
        ws_deps.cell(row=row_num, column=1, value=dep.environment_name)
        ws_deps.cell(row=row_num, column=2, value=dep.depends_on_name)
        ws_deps.cell(row=row_num, column=3, value=dep.dependency_type)

    # Auto-size columns
    for col_num in range(1, len(dep_headers) + 1):
        ws_deps.column_dimensions[get_column_letter(col_num)].width = 25

    # Sheet 3: Placement (empty for now - future implementation)
    ws_placement = wb.create_sheet(title="Placement")
    placement_headers = ['environment_name', 'primary_node', 'secondary_node', 'recorded_at']
    for col_num, header in enumerate(placement_headers, 1):
        cell = ws_placement.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Note: Placement history will be populated when poller is implemented
    ws_placement.cell(row=2, column=1, value="Placement history not yet available")

    for col_num in range(1, len(placement_headers) + 1):
        ws_placement.column_dimensions[get_column_letter(col_num)].width = 20

    # Save to BytesIO and return as response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="environments.xlsx"'

    return response


def blast_radius(request, name):
    """
    API endpoint: GET /api/environments/<name>/blast-radius/

    Returns all environments that would be affected if the given environment goes down.
    Uses recursive CTE to traverse the dependency graph.
    """
    # Verify environment exists
    env = get_object_or_404(Environment, name=name)

    # Recursive CTE query to find all dependent environments
    with connection.cursor() as cursor:
        cursor.execute("""
            WITH RECURSIVE blast_radius AS (
                -- Base case: direct dependents of the target environment
                SELECT
                    ed.environment_name,
                    ed.depends_on_name,
                    ed.dependency_type,
                    1 as depth
                FROM environment_dependencies ed
                WHERE ed.depends_on_name = %s

                UNION

                -- Recursive case: dependents of dependents
                SELECT
                    ed.environment_name,
                    ed.depends_on_name,
                    ed.dependency_type,
                    br.depth + 1
                FROM environment_dependencies ed
                INNER JOIN blast_radius br ON ed.depends_on_name = br.environment_name
                WHERE br.depth < 10  -- Prevent infinite loops
            )
            SELECT DISTINCT
                br.environment_name,
                br.depth,
                br.dependency_type,
                e.env_type,
                e.status,
                e.criticality_tier,
                e.owner,
                e.team,
                e.region
            FROM blast_radius br
            LEFT JOIN environments e ON e.name = br.environment_name
            ORDER BY br.depth, br.environment_name
        """, [name])

        columns = [col[0] for col in cursor.description]
        results = []
        for row in cursor.fetchall():
            results.append(dict(zip(columns, row)))

    # Storage co-accessors (#46): environments that share a StorageResource with
    # the target are indirect blast-radius members (dependency_type='storage').
    from cmdb.apps.storage.models import StorageEnvironmentAccess
    existing = {r['environment_name'] for r in results}
    storage_ids = list(
        StorageEnvironmentAccess.objects.filter(environment=env).values_list('storage_id', flat=True)
    )
    storage_results = []
    if storage_ids:
        co_names = set(
            StorageEnvironmentAccess.objects.filter(storage_id__in=storage_ids)
            .exclude(environment=env)
            .values_list('environment__name', flat=True)
        ) - existing - {name}
        attrs = {e.name: e for e in Environment.objects.filter(name__in=co_names)}
        for n in sorted(co_names):
            e = attrs.get(n)
            storage_results.append({
                'environment_name': n, 'depth': 1, 'dependency_type': 'storage',
                'env_type': getattr(e, 'env_type', None), 'status': getattr(e, 'status', None),
                'criticality_tier': getattr(e, 'criticality_tier', None),
                'owner': getattr(e, 'owner', None), 'team': getattr(e, 'team', None),
                'region': getattr(e, 'region', None),
            })
    results.extend(storage_results)

    return JsonResponse({
        'target': name,
        'affected_count': len(results),
        'storage_shared_count': len(storage_results),
        'affected_environments': results
    })


def charm_statistics(request):
    """
    View showing charm usage statistics with counts per charm and version.
    """
    from collections import defaultdict

    # Collect charm statistics
    charm_stats = defaultdict(lambda: {'total': 0, 'versions': defaultdict(list)})

    for env in Environment.objects.exclude(charm_versions={}):
        for charm_name, version in env.charm_versions.items():
            charm_stats[charm_name]['total'] += 1
            charm_stats[charm_name]['versions'][version].append({
                'name': env.name,
                'env_type': env.env_type,
                'region': env.region,
                'owner': env.owner,
            })

    # Convert to sorted list
    stats_list = []
    for charm_name, data in sorted(charm_stats.items()):
        versions = []
        for version, envs in sorted(data['versions'].items()):
            versions.append({
                'version': version,
                'count': len(envs),
                'environments': sorted(envs, key=lambda x: x['name'])
            })
        stats_list.append({
            'charm': charm_name,
            'total_environments': data['total'],
            'versions': versions
        })

    context = {
        'charm_stats': stats_list,
        'total_charms': len(stats_list),
        'total_deployments': sum(s['total_environments'] for s in stats_list),
        'active_tab': 'overview',
    }

    return render(request, 'environments/charm_stats.html', context)


def charm_outdated(request):
    """
    Charms tab → "Outdated" subtab.

    Cross-checks each environment's declared charm_versions against the cached
    Charmhub catalogue (CharmRelease, refreshed by `manage.py refresh_charmhub`)
    and lists, per Juju model, the charms that are either behind the latest
    published revision or — for prod — not running on the `stable` risk.

    Reads only the database (charm_versions + the Charmhub cache); it never
    touches the infra/terraform repos.
    """
    from .charm_utils import parse_charm_version, track_tuple
    from .models import CharmRelease

    prod_only = request.GET.get('prod') == '1'

    # Charmhub cache as a {(charm, track, risk): CharmRelease} lookup.
    catalog = {(r.charm, r.track, r.risk): r for r in CharmRelease.objects.all()}
    checked_at = max((r.checked_at for r in catalog.values()), default=None)

    # Newest *numeric* stable track per charm, e.g. postgresql -> (16,) on
    # 16/stable. Used to flag charms pinned to an older major track.
    newest_stable: dict[str, tuple[tuple[int, ...], CharmRelease]] = {}
    for (charm, track, risk), rel in catalog.items():
        if risk != 'stable':
            continue
        tt = track_tuple(track)
        if tt is None:
            continue
        cur = newest_stable.get(charm)
        if cur is None or tt > cur[0]:
            newest_stable[charm] = (tt, rel)

    def fmt(rel):
        if rel is None:
            return '—'
        if rel.latest_revision is not None:
            return f"{rel.track}/{rel.risk} (rev {rel.latest_revision})"
        return f"{rel.track}/{rel.risk}"

    rows = []
    for env in Environment.objects.exclude(charm_versions={}).order_by('name'):
        is_prod = env.env_type == 'prod'
        if prod_only and not is_prod:
            continue
        for charm_name, value in env.charm_versions.items():
            parsed = parse_charm_version(value)
            if parsed is None or not parsed.has_channel:
                continue  # placeholder / revision-only: nothing to compare

            latest = catalog.get((charm_name, parsed.track, parsed.risk))
            reasons = []
            target = None
            behind_by = None
            newer_track = None

            # Hard rule: prod must be on stable. This dominates — when the
            # channel itself is wrong, "move to stable" is the action, so we
            # don't also compare revisions within the wrong channel (stable
            # legitimately lags edge, which would read as a confusing
            # "N revisions behind" against a lower target revision).
            if is_prod and parsed.risk != 'stable':
                reasons.append('prod not on stable')
                target = catalog.get((charm_name, parsed.track, 'stable'))

            # Otherwise: a pinned revision behind the latest in its own channel.
            elif (parsed.revision is not None and latest is not None
                    and latest.latest_revision is not None
                    and parsed.revision < latest.latest_revision):
                reasons.append('revision behind')
                behind_by = latest.latest_revision - parsed.revision
                target = latest

            # Independent signal: a newer numeric stable track exists, e.g. the
            # charm is on 14/stable while 16/stable ships. (Major-track moves are
            # migrations, so this is advisory and shown alongside any of the above.)
            cur_tt = track_tuple(parsed.track)
            nst = newest_stable.get(charm_name)
            if cur_tt is not None and nst is not None and nst[0] > cur_tt:
                reasons.append('newer track')
                newer_track = fmt(nst[1])
                if target is None:
                    target = nst[1]

            if not reasons:
                continue

            rows.append({
                'model': env.name,
                'env_type': env.env_type,
                'is_prod': is_prod,
                'cloud': env.cloud,
                'region': env.region,
                'owner': env.owner,
                'charm': charm_name,
                'current': value,
                'latest': fmt(target),
                'on_charmhub': latest is not None or target is not None,
                'reasons': reasons,
                'behind_by': behind_by,
                'newer_track': newer_track,
            })

    # Prod first, then by charm, then model.
    rows.sort(key=lambda r: (0 if r['is_prod'] else 1, r['charm'], r['model']))

    context = {
        'rows': rows,
        'total_flagged': len(rows),
        'prod_not_stable': sum(1 for r in rows if 'prod not on stable' in r['reasons']),
        'revision_behind': sum(1 for r in rows if 'revision behind' in r['reasons']),
        'newer_track': sum(1 for r in rows if 'newer track' in r['reasons']),
        'models_affected': len({r['model'] for r in rows}),
        'checked_at': checked_at,
        'cache_empty': not catalog,
        'prod_only': prod_only,
        'active_tab': 'outdated',
    }
    return render(request, 'environments/charm_outdated.html', context)


def cia_assessment(request):
    """
    CIA assessment dashboard showing risk profiles and compliance scope.
    """
    from collections import defaultdict
    from django.db.models import Count, Q

    # Overall statistics
    total_envs = Environment.objects.count()

    # By criticality tier
    tier_stats = []
    for tier in [1, 2, 3]:
        envs = Environment.objects.filter(criticality_tier=tier)
        tier_stats.append({
            'tier': tier,
            'name': {1: 'Critical', 2: 'Important', 3: 'Best Effort'}[tier],
            'count': envs.count(),
            'prod_count': envs.filter(env_type='prod').count(),
            'environments': list(envs.values('name', 'owner', 'region', 'env_type', 'status')[:100])
        })

    # By data classification
    data_class_stats = []
    for dc in ['pii', 'internal', 'public']:
        envs = Environment.objects.filter(data_classification=dc)
        data_class_stats.append({
            'classification': dc.upper(),
            'count': envs.count(),
            'critical_count': envs.filter(criticality_tier=1).count(),
            'environments': list(envs.values('name', 'owner', 'criticality_tier', 'env_type')[:100])
        })

    # By SLO level
    slo_stats = Environment.objects.filter(slo_level__isnull=False)\
        .values('slo_level')\
        .annotate(count=Count('id'))\
        .order_by('-count')

    # Compliance scope analysis
    compliance_envs = Environment.objects.exclude(compliance_scope=[])
    compliance_counts = defaultdict(int)
    compliance_envs_by_scope = defaultdict(list)

    for env in compliance_envs:
        if isinstance(env.compliance_scope, list):
            for scope in env.compliance_scope:
                compliance_counts[scope] += 1
                compliance_envs_by_scope[scope].append({
                    'name': env.name,
                    'owner': env.owner,
                    'criticality_tier': env.criticality_tier,
                })

    compliance_stats = [
        {
            'scope': scope,
            'count': count,
            'environments': compliance_envs_by_scope[scope][:50]
        }
        for scope, count in sorted(compliance_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    # High-risk environments (Tier 1 + PII + Production)
    high_risk = Environment.objects.filter(
        criticality_tier=1,
        data_classification='pii',
        env_type='prod',
        status='active'
    ).values('name', 'owner', 'cia_owner', 'slo_level', 'region')[:50]

    # Environments missing CIA data
    missing_cia = Environment.objects.filter(
        Q(criticality_tier__isnull=True) |
        Q(data_classification__isnull=True) |
        Q(cia_owner__isnull=True)
    ).filter(status='active').count()

    context = {
        'total_envs': total_envs,
        'tier_stats': tier_stats,
        'data_class_stats': data_class_stats,
        'slo_stats': slo_stats,
        'compliance_stats': compliance_stats,
        'high_risk': list(high_risk),
        'missing_cia': missing_cia,
    }

    return render(request, 'environments/cia_assessment.html', context)


def team_aggregation(request):
    """
    Team resource aggregation view showing quota usage by team.
    Supports filtering by region, cloud, and architecture.
    """
    from django.db.models import Count, Sum, Q, Case, When, IntegerField
    from collections import defaultdict

    # Get filter parameters
    region_filter = request.GET.get('region')
    cloud_filter = request.GET.get('cloud')
    arch_filter = request.GET.get('architecture')

    # Base queryset
    queryset = Environment.objects.all()

    # Apply filters
    if region_filter and region_filter != 'all':
        queryset = queryset.filter(region=region_filter)
    if cloud_filter and cloud_filter != 'all':
        queryset = queryset.filter(cloud=cloud_filter)
    if arch_filter and arch_filter != 'all':
        queryset = queryset.filter(compute_architecture=arch_filter)

    # Aggregate by team
    team_stats = defaultdict(lambda: {
        'total_envs': 0,
        'cores': 0,
        'ram': 0,
        'instances': 0,
        'gigabytes': 0,
        'volumes': 0,
        'by_region': defaultdict(int),
        'by_cloud': defaultdict(int),
        'by_env_type': defaultdict(int),
    })

    for env in queryset.select_related():
        team = env.team or 'Unknown'
        team_stats[team]['total_envs'] += 1

        # Aggregate quotas
        if env.quotas:
            team_stats[team]['cores'] += env.quotas.get('cores', 0) or 0
            team_stats[team]['ram'] += env.quotas.get('ram', 0) or 0
            team_stats[team]['instances'] += env.quotas.get('instances', 0) or 0
            team_stats[team]['gigabytes'] += env.quotas.get('gigabytes', 0) or 0
            team_stats[team]['volumes'] += env.quotas.get('volumes', 0) or 0

        # Breakdown by region, cloud, env_type
        if env.region:
            team_stats[team]['by_region'][env.region] += 1
        if env.cloud:
            team_stats[team]['by_cloud'][env.cloud] += 1
        if env.env_type:
            team_stats[team]['by_env_type'][env.env_type] += 1

    # Convert to sorted list
    teams_list = []
    for team, stats in sorted(team_stats.items()):
        teams_list.append({
            'team': team,
            'total_envs': stats['total_envs'],
            'cores': stats['cores'],
            'ram_gb': stats['ram'] / 1024 if stats['ram'] else 0,
            'instances': stats['instances'],
            'gigabytes': stats['gigabytes'],
            'volumes': stats['volumes'],
            'by_region': dict(stats['by_region']),
            'by_cloud': dict(stats['by_cloud']),
            'by_env_type': dict(stats['by_env_type']),
            # Add display keys for table columns
            'regions_display': dict(stats['by_region']),
            'clouds_display': dict(stats['by_cloud']),
            'types_display': dict(stats['by_env_type']),
        })

    # Carry the active filters into each team's drill-down link.
    _eq = []
    if region_filter and region_filter != 'all':
        _eq.append(f'region={region_filter}')
    if cloud_filter and cloud_filter != 'all':
        _eq.append(f'cloud={cloud_filter}')
    if arch_filter and arch_filter != 'all':
        _eq.append(f'compute_architecture={arch_filter}')
    extra_query = ('&' + '&'.join(_eq)) if _eq else ''

    # Create table and configure for sorting/pagination
    table = TeamAggregationTable(teams_list, extra_query=extra_query)
    RequestConfig(request, paginate={'per_page': 50}).configure(table)

    # Get filter choices
    regions = [
        {'value': 'all', 'name': 'All Regions'},
        {'value': 'amer', 'name': 'AMER'},
        {'value': 'emea', 'name': 'EMEA'},
        {'value': 'apac', 'name': 'APAC'},
    ]

    clouds = [{'value': 'all', 'name': 'All Clouds'}]
    clouds.extend([
        {'value': c, 'name': c}
        for c in Environment.objects.values_list('cloud', flat=True).distinct().order_by('cloud')
        if c
    ])

    architectures = [{'value': 'all', 'name': 'All Architectures'}]
    architectures.extend([
        {'value': arch, 'name': arch}
        for arch in Environment.objects.values_list('compute_architecture', flat=True).distinct().order_by('compute_architecture')
        if arch
    ])

    # Chart data — derived from the quota aggregation we already computed, so the
    # charts don't depend on the (currently empty) CloudCapacity / Redis path.
    # Top 20 teams by cores keeps the bar charts legible.
    chart_data = [
        {
            'team': t['team'],
            'cores': t['cores'],
            'ram_gb': round(t['ram_gb'], 1),
            'env_count': t['total_envs'],
        }
        for t in sorted(teams_list, key=lambda t: t['cores'], reverse=True)[:20]
    ]

    context = {
        'table': table,
        'teams': teams_list,
        'chart_data': chart_data,
        'total_teams': len(teams_list),
        'total_envs': sum(t['total_envs'] for t in teams_list),
        'total_cores': sum(t['cores'] for t in teams_list),
        'total_ram_gb': sum(t['ram_gb'] for t in teams_list),
        'total_instances': sum(t['instances'] for t in teams_list),
        'regions': regions,
        'clouds': clouds,
        'architectures': architectures,
        'active_region': region_filter or 'all',
        'active_cloud': cloud_filter or 'all',
        'active_architecture': arch_filter or 'all',
    }

    return render(request, 'environments/team_aggregation.html', context)


def autocomplete(request):
    """
    API endpoint for autocomplete on various fields.
    Returns matching results based on field type and query parameter.
    """
    query = request.GET.get('q', '').strip()
    field = request.GET.get('field', 'name')  # name, owner, team, cia_owner, cia_custodian

    if len(query) < 2:
        return JsonResponse({'results': []})

    results = []

    if field == 'name':
        # Search environment names
        envs = Environment.objects.filter(name__icontains=query)\
            .values('name', 'owner', 'env_type', 'region')[:20]
        results = [
            {
                'value': env['name'],
                'label': f"{env['name']} ({env['env_type']}, {env['region'] or '?'}, {env['owner'] or 'no owner'})"
            }
            for env in envs
        ]

    elif field in ['owner', 'team']:
        # Get unique owners/teams
        values = Environment.objects.filter(
            **{f'{field}__icontains': query}
        ).values_list(field, flat=True).distinct().order_by(field)[:20]
        results = [{'value': v, 'label': v} for v in values if v]

    elif field == 'cia_owner':
        # Get unique CIA owners
        values = Environment.objects.filter(cia_owner__icontains=query)\
            .values_list('cia_owner', flat=True).distinct().order_by('cia_owner')[:20]
        results = [{'value': v, 'label': v} for v in values if v]

    elif field == 'cia_custodian':
        # Get unique CIA custodians
        values = Environment.objects.filter(cia_custodian__icontains=query)\
            .values_list('cia_custodian', flat=True).distinct().order_by('cia_custodian')[:20]
        results = [{'value': v, 'label': v} for v in values if v]

    elif field == 'charm':
        # Get unique charm names from charm_versions JSON field
        from collections import Counter
        charm_counts = Counter()
        for env in Environment.objects.exclude(charm_versions={}):
            for charm_name in env.charm_versions.keys():
                if query.lower() in charm_name.lower():
                    charm_counts[charm_name] += 1

        # Return top 20 matching charms sorted by usage
        results = [
            {
                'value': charm,
                'label': f"{charm} ({count} environments)"
            }
            for charm, count in charm_counts.most_common(20)
        ]

    return JsonResponse({'results': results})


def cloud_region_capacity(request):
    """
    Cloud/Region capacity view showing resource distribution across clouds and regions.
    """
    from collections import defaultdict

    # Get filter parameters
    cloud_filter = request.GET.get('cloud')
    region_filter = request.GET.get('region')

    # Base queryset
    queryset = Environment.objects.all()

    # Apply filters
    if cloud_filter and cloud_filter != 'all':
        queryset = queryset.filter(cloud=cloud_filter)
    if region_filter and region_filter != 'all':
        queryset = queryset.filter(region=region_filter)

    # Aggregate by cloud + region
    cloud_region_stats = defaultdict(lambda: {
        'cloud': '',
        'region': '',
        'total_envs': 0,
        'cores': 0,
        'ram': 0,
        'instances': 0,
        'gigabytes': 0,
        'volumes': 0,
        'prod_count': 0,
        'staging_count': 0,
        'dev_count': 0,
    })

    for env in queryset.select_related():
        key = (env.cloud or 'Unknown', env.region or 'Unknown')
        stats = cloud_region_stats[key]
        stats['cloud'] = env.cloud or 'Unknown'
        stats['region'] = env.region or 'Unknown'
        stats['total_envs'] += 1

        # Aggregate quotas
        if env.quotas:
            stats['cores'] += env.quotas.get('cores', 0) or 0
            stats['ram'] += env.quotas.get('ram', 0) or 0
            stats['instances'] += env.quotas.get('instances', 0) or 0
            stats['gigabytes'] += env.quotas.get('gigabytes', 0) or 0
            stats['volumes'] += env.quotas.get('volumes', 0) or 0

        # Count by env_type
        if env.env_type == 'prod':
            stats['prod_count'] += 1
        elif env.env_type == 'staging':
            stats['staging_count'] += 1
        elif env.env_type == 'dev':
            stats['dev_count'] += 1

    # Convert to sorted list
    cloud_region_list = []
    for key, stats in sorted(cloud_region_stats.items()):
        cloud_region_list.append({
            'cloud_region': f"{stats['cloud']} / {stats['region']}",
            'cloud': stats['cloud'],
            'region': stats['region'],
            'total_envs': stats['total_envs'],
            'cores': stats['cores'],
            'ram_gb': stats['ram'] / 1024 if stats['ram'] else 0,
            'instances': stats['instances'],
            'gigabytes': stats['gigabytes'],
            'volumes': stats['volumes'],
            'prod_count': stats['prod_count'],
            'staging_count': stats['staging_count'],
        })

    # Create table and configure for sorting/pagination
    table = CloudRegionTable(cloud_region_list)
    RequestConfig(request, paginate={'per_page': 50}).configure(table)

    # Get filter choices
    regions = [{'value': 'all', 'name': 'All Regions'}]
    regions.extend([
        {'value': r, 'name': r.upper()}
        for r in Environment.objects.values_list('region', flat=True).distinct().order_by('region')
        if r
    ])

    clouds = [{'value': 'all', 'name': 'All Clouds'}]
    clouds.extend([
        {'value': c, 'name': c}
        for c in Environment.objects.values_list('cloud', flat=True).distinct().order_by('cloud')
        if c
    ])

    context = {
        'table': table,
        'cloud_regions': cloud_region_list,
        'total_cloud_regions': len(cloud_region_list),
        'total_envs': sum(cr['total_envs'] for cr in cloud_region_list),
        'total_cores': sum(cr['cores'] for cr in cloud_region_list),
        'total_ram_gb': sum(cr['ram_gb'] for cr in cloud_region_list),
        'regions': regions,
        'clouds': clouds,
        'active_region': region_filter or 'all',
        'active_cloud': cloud_filter or 'all',
    }

    return render(request, 'environments/cloud_region_capacity.html', context)


def controller_health(request):
    """
    Controller health dashboard showing environments grouped by Juju controller.
    """
    from collections import defaultdict

    # Get filter parameters
    stage_filter = request.GET.get('stage')

    # Base queryset
    queryset = Environment.objects.all()

    # Apply stage filter
    if stage_filter and stage_filter != 'all':
        queryset = queryset.filter(juju_controller_stage=stage_filter)

    from collections import Counter

    # Aggregate by controller + stage
    controller_stats = defaultdict(lambda: {
        'juju_controller': '',
        'stage': '',
        'total_envs': 0,
        'active_count': 0,
        'degraded_count': 0,
        'maintenance_count': 0,
        'tier1_count': 0,
        'tier2_count': 0,
        'versions': Counter(),
        'clouds': Counter(),
        'environments': [],
    })

    for env in queryset.select_related():
        controller = env.juju_controller or 'Unknown'
        stage = env.juju_controller_stage or 'Unknown'
        key = (controller, stage)
        stats = controller_stats[key]
        stats['juju_controller'] = controller
        stats['stage'] = stage
        stats['total_envs'] += 1
        if env.juju_series:
            stats['versions'][env.juju_series] += 1
        if env.cloud:
            stats['clouds'][env.cloud] += 1

        # Count by status
        if env.status == 'active':
            stats['active_count'] += 1
        elif env.status == 'degraded':
            stats['degraded_count'] += 1
        elif env.status == 'maintenance':
            stats['maintenance_count'] += 1

        # Count by criticality tier
        if env.criticality_tier == 1:
            stats['tier1_count'] += 1
        elif env.criticality_tier == 2:
            stats['tier2_count'] += 1

        stats['environments'].append(env.name)

    # Calculate blast radius per controller using recursive CTE
    with connection.cursor() as cursor:
        for key, stats in controller_stats.items():
            env_names = stats['environments']
            if not env_names:
                stats['depended_on_count'] = 0
                continue

            # Calculate total blast radius for all environments in this controller
            cursor.execute("""
                WITH RECURSIVE blast_radius AS (
                    SELECT DISTINCT
                        ed.environment_name
                    FROM environment_dependencies ed
                    WHERE ed.depends_on_name = ANY(%s)

                    UNION

                    SELECT DISTINCT
                        ed.environment_name
                    FROM environment_dependencies ed
                    INNER JOIN blast_radius br ON ed.depends_on_name = br.environment_name
                )
                SELECT COUNT(DISTINCT environment_name) FROM blast_radius
            """, [env_names])

            stats['depended_on_count'] = cursor.fetchone()[0]

    # Convert to sorted list
    controller_list = []
    for key, stats in sorted(controller_stats.items(), key=lambda x: x[1]['total_envs'], reverse=True):
        controller_list.append({
            'juju_controller': stats['juju_controller'],
            'stage': stats['stage'],
            'version': stats['versions'].most_common(1)[0][0] if stats['versions'] else '—',
            'cloud': ', '.join(c for c, _ in stats['clouds'].most_common()) or '—',
            'total_envs': stats['total_envs'],
            'active_count': stats['active_count'],
            'degraded_count': stats['degraded_count'],
            'maintenance_count': stats['maintenance_count'],
            'tier1_count': stats['tier1_count'],
            'tier2_count': stats['tier2_count'],
            'depended_on_count': stats['depended_on_count'],
        })

    # Create table and configure for sorting/pagination
    table = ControllerTable(controller_list)
    RequestConfig(request, paginate={'per_page': 50}).configure(table)

    # Get stage choices
    stages = [{'value': 'all', 'name': 'All Stages'}]
    stages.extend([
        {'value': s, 'name': s}
        for s in Environment.objects.values_list('juju_controller_stage', flat=True).distinct().order_by('juju_controller_stage')
        if s
    ])

    context = {
        'table': table,
        'controllers': controller_list,
        'total_controllers': len(controller_list),
        'total_envs': sum(c['total_envs'] for c in controller_list),
        'total_degraded': sum(c['degraded_count'] for c in controller_list),
        'stages': stages,
        'active_stage': stage_filter or 'all',
    }

    return render(request, 'environments/controller_health.html', context)


def dependency_hotspots(request):
    """
    Dependency hotspots view ranking environments by how many others depend on them.
    Shows reverse blast radius (most critical dependencies).
    """
    # Calculate direct dependents count and total blast radius for each environment
    # First, get all environments that are depended upon
    depended_on_envs = EnvironmentDependency.objects.values_list('depends_on_name', flat=True).distinct()

    hotspots = []
    with connection.cursor() as cursor:
        for env_name in depended_on_envs:
            # Count direct dependents
            direct_count = EnvironmentDependency.objects.filter(depends_on_name=env_name).count()

            # Calculate total blast radius using recursive CTE
            cursor.execute("""
                WITH RECURSIVE blast_radius AS (
                    SELECT DISTINCT environment_name
                    FROM environment_dependencies
                    WHERE depends_on_name = %s

                    UNION

                    SELECT DISTINCT ed.environment_name
                    FROM environment_dependencies ed
                    INNER JOIN blast_radius br ON ed.depends_on_name = br.environment_name
                )
                SELECT COUNT(DISTINCT environment_name) FROM blast_radius
            """, [env_name])

            total_blast = cursor.fetchone()[0]

            # Get environment details
            try:
                env = Environment.objects.get(name=env_name)
                hotspots.append({
                    'environment_name': env.name,
                    'depended_on_count': direct_count,
                    'total_blast_radius': total_blast,
                    'criticality_tier': env.criticality_tier,
                    'env_type': env.env_type,
                    'status': env.status,
                    'owner': env.owner,
                    'team': env.team,
                    'region': env.region,
                })
            except Environment.DoesNotExist:
                # Environment might have been deleted but dependency record remains
                hotspots.append({
                    'environment_name': env_name,
                    'depended_on_count': direct_count,
                    'total_blast_radius': total_blast,
                    'criticality_tier': None,
                    'env_type': None,
                    'status': None,
                    'owner': None,
                    'team': None,
                    'region': None,
                })

    # Sort by blast radius and then by direct count
    hotspots.sort(key=lambda x: (x['total_blast_radius'], x['depended_on_count']), reverse=True)


    # Create table and configure for sorting/pagination
    table = DependencyHotspotTable(hotspots)
    RequestConfig(request, paginate={'per_page': 50}).configure(table)

    context = {
        'table': table,
        'hotspots': hotspots[:20],  # Top 20 for summary
        'total_hotspots': len(hotspots),
    }

    return render(request, 'environments/dependency_hotspots.html', context)


def version_compliance(request):
    """
    Version compliance view showing distribution of PostgreSQL versions,
    Juju series, and charm versions across environments.
    """
    from collections import Counter

    # PostgreSQL versions
    pg_versions = Environment.objects.exclude(postgresql_major_version__isnull=True)\
        .values('postgresql_major_version')\
        .annotate(
            count=Count('id'),
            tier1_count=Count('id', filter=Q(criticality_tier=1)),
            prod_count=Count('id', filter=Q(env_type='prod'))
        ).order_by('-count')

    # Juju series
    juju_series = Environment.objects.exclude(juju_series__isnull=True)\
        .values('juju_series')\
        .annotate(
            count=Count('id'),
            tier1_count=Count('id', filter=Q(criticality_tier=1)),
            prod_count=Count('id', filter=Q(env_type='prod'))
        ).order_by('-count')

    # Compute architecture distribution
    compute_arch = Environment.objects.exclude(compute_architecture__isnull=True)\
        .values('compute_architecture')\
        .annotate(
            count=Count('id'),
            tier1_count=Count('id', filter=Q(criticality_tier=1)),
            prod_count=Count('id', filter=Q(env_type='prod'))
        ).order_by('-count')

    # Charm version statistics (top 20 charms by usage)
    charm_counts = Counter()
    for env in Environment.objects.exclude(charm_versions={}):
        for charm_name in env.charm_versions.keys():
            charm_counts[charm_name] += 1

    top_charms = charm_counts.most_common(20)

    # Build version compliance data for table
    version_data = []

    # Add PostgreSQL versions
    for pv in pg_versions:
        version_data.append({
            'version_type': 'PostgreSQL',
            'version_value': pv['postgresql_major_version'] or 'Unknown',
            'count': pv['count'],
            'tier1_count': pv['tier1_count'],
            'prod_count': pv['prod_count'],
        })

    # Add Juju series
    for js in juju_series:
        version_data.append({
            'version_type': 'Juju Series',
            'version_value': js['juju_series'] or 'Unknown',
            'count': js['count'],
            'tier1_count': js['tier1_count'],
            'prod_count': js['prod_count'],
        })

    # Add compute architecture
    for ca in compute_arch:
        version_data.append({
            'version_type': 'Compute Arch',
            'version_value': ca['compute_architecture'] or 'Unknown',
            'count': ca['count'],
            'tier1_count': ca['tier1_count'],
            'prod_count': ca['prod_count'],
        })

    # Create table and configure for sorting/pagination
    table = VersionComplianceTable(version_data)
    RequestConfig(request, paginate={'per_page': 100}).configure(table)

    context = {
        'table': table,
        'pg_versions': list(pg_versions),
        'juju_series': list(juju_series),
        'compute_arch': list(compute_arch),
        'top_charms': top_charms,
        'total_pg_versions': len(pg_versions),
        'total_juju_series': len(juju_series),
    }

    return render(request, 'environments/version_compliance.html', context)


def owner_dashboard(request):
    """
    Owner dashboard showing personalized view of environments owned by a specific person.
    If no owner specified, shows the current user's environments (from query param or auth).
    """
    # Get owner from query param
    owner = request.GET.get('owner', '')

    if not owner:
        # In production, get from authenticated user
        # For now, show a selection page
        return render(request, 'environments/owner_dashboard_select.html', {
            'owners': Environment.objects.values_list('owner', flat=True)
                      .distinct().order_by('owner').exclude(owner__isnull=True)
        })

    # Get environments for this owner
    envs = Environment.objects.filter(owner=owner)

    # Overall stats
    total_envs = envs.count()
    prod_count = envs.filter(env_type='prod').count()
    tier1_count = envs.filter(criticality_tier=1).count()
    tier2_count = envs.filter(criticality_tier=2).count()

    # Status breakdown
    active_count = envs.filter(status='active').count()
    degraded_count = envs.filter(status='degraded').count()
    maintenance_count = envs.filter(status='maintenance').count()

    # Quota totals
    total_cores = 0
    total_ram = 0
    total_instances = 0
    for env in envs:
        if env.quotas:
            total_cores += env.quotas.get('cores', 0) or 0
            total_ram += env.quotas.get('ram', 0) or 0
            total_instances += env.quotas.get('instances', 0) or 0

    # Environments needing attention
    needs_attention = []

    # Missing CIA data
    missing_cia = envs.filter(
        Q(criticality_tier__isnull=True) |
        Q(data_classification__isnull=True) |
        Q(cia_owner__isnull=True)
    ).filter(status='active').values('name', 'env_type', 'region')

    for env in missing_cia:
        needs_attention.append({
            'name': env['name'],
            'reason': 'Missing CIA data',
            'severity': 'warning'
        })

    # Degraded environments
    degraded = envs.filter(status='degraded').values('name', 'env_type', 'region')
    for env in degraded:
        needs_attention.append({
            'name': env['name'],
            'reason': 'Degraded status',
            'severity': 'danger'
        })

    # Missing runbooks for Tier 1 prod
    missing_runbooks = envs.filter(
        criticality_tier=1,
        env_type='prod',
        runbook_url__isnull=True
    ).values('name', 'region')

    for env in missing_runbooks:
        needs_attention.append({
            'name': env['name'],
            'reason': 'Tier 1 prod missing runbook',
            'severity': 'warning'
        })

    # Region distribution
    region_dist = {}
    for env in envs:
        region = env.region or 'Unknown'
        region_dist[region] = region_dist.get(region, 0) + 1

    context = {
        'owner': owner,
        'total_envs': total_envs,
        'prod_count': prod_count,
        'tier1_count': tier1_count,
        'tier2_count': tier2_count,
        'active_count': active_count,
        'degraded_count': degraded_count,
        'maintenance_count': maintenance_count,
        'total_cores': total_cores,
        'total_ram_gb': total_ram / 1024 if total_ram else 0,
        'total_instances': total_instances,
        'needs_attention': needs_attention[:20],
        'region_dist': region_dist,
        'environments': envs.values(
            'name', 'env_type', 'status', 'criticality_tier', 'region', 'team'
        ).order_by('-criticality_tier', 'name')[:50],
    }

    return render(request, 'environments/owner_dashboard.html', context)


def lifecycle_timeline(request):
    """
    Lifecycle timeline showing environment creation/decommissioning trends and age distribution.
    """
    from django.db.models import Count
    from django.db.models.functions import TruncMonth, TruncDate
    from datetime import datetime, timedelta

    # Environments created per month (last 12 months)
    twelve_months_ago = datetime.now() - timedelta(days=365)
    created_by_month = Environment.objects.filter(created_at__gte=twelve_months_ago)\
        .annotate(month=TruncMonth('created_at'))\
        .values('month')\
        .annotate(count=Count('id'))\
        .order_by('month')

    # Environments decommissioned (with end_date)
    decommissioned = Environment.objects.exclude(end_date__isnull=True)\
        .filter(end_date__gte=twelve_months_ago)\
        .annotate(month=TruncMonth('end_date'))\
        .values('month')\
        .annotate(count=Count('id'))\
        .order_by('month')

    # Age distribution (in days)
    from django.utils import timezone
    now = timezone.now()

    age_buckets = {
        '0-30 days': 0,
        '31-90 days': 0,
        '91-180 days': 0,
        '181-365 days': 0,
        '1-2 years': 0,
        '2+ years': 0,
    }

    for env in Environment.objects.filter(end_date__isnull=True):
        if env.created_at:
            age_days = (now - env.created_at).days
            if age_days <= 30:
                age_buckets['0-30 days'] += 1
            elif age_days <= 90:
                age_buckets['31-90 days'] += 1
            elif age_days <= 180:
                age_buckets['91-180 days'] += 1
            elif age_days <= 365:
                age_buckets['181-365 days'] += 1
            elif age_days <= 730:
                age_buckets['1-2 years'] += 1
            else:
                age_buckets['2+ years'] += 1

    # Environments in decommissioning pipeline
    decommissioning = Environment.objects.filter(status='decommissioning')\
        .values('name', 'owner', 'team', 'region', 'env_type', 'end_date')\
        .order_by('end_date')

    # Newest and oldest environments
    newest = Environment.objects.filter(end_date__isnull=True)\
        .order_by('-created_at')\
        .values('name', 'created_at', 'owner', 'env_type', 'region')[:10]

    oldest = Environment.objects.filter(end_date__isnull=True)\
        .order_by('created_at')\
        .values('name', 'created_at', 'owner', 'env_type', 'region')[:10]

    context = {
        'created_by_month': list(created_by_month),
        'decommissioned': list(decommissioned),
        'age_buckets': age_buckets,
        'decommissioning': list(decommissioning)[:50],
        'decommissioning_count': len(decommissioning),
        'newest': list(newest),
        'oldest': list(oldest),
        'total_active': Environment.objects.filter(end_date__isnull=True).count(),
        'total_decommissioned': Environment.objects.exclude(end_date__isnull=True).count(),
    }

    return render(request, 'environments/lifecycle_timeline.html', context)


def risk_heatmap(request):
    """
    Risk heatmap showing matrix of criticality_tier vs data_classification.
    Helps identify high-risk environment clusters.
    """
    from collections import defaultdict

    # Build matrix: tier x classification
    matrix = defaultdict(lambda: defaultdict(list))

    for env in Environment.objects.filter(status='active'):
        tier = env.criticality_tier if env.criticality_tier else 'Unknown'
        classification = env.data_classification or 'Unknown'
        matrix[tier][classification].append({
            'name': env.name,
            'owner': env.owner,
            'env_type': env.env_type,
            'region': env.region,
            'runbook_url': env.runbook_url,
        })

    # Convert to structured data for template
    tiers = [1, 2, 3, 'Unknown']
    classifications = ['pii', 'internal', 'public', 'Unknown']

    heatmap_data = []
    for tier in tiers:
        row = {'tier': tier, 'cells': []}
        for classification in classifications:
            envs = matrix[tier][classification]
            row['cells'].append({
                'classification': classification,
                'count': len(envs),
                'environments': envs[:20],  # Limit to 20 for display
            })
        heatmap_data.append(row)

    # High-risk segments (Tier 1 + PII)
    high_risk = matrix[1]['pii']
    high_risk_no_runbook = [e for e in high_risk if not e['runbook_url']]

    # Missing classification data
    missing_data = Environment.objects.filter(
        Q(criticality_tier__isnull=True) | Q(data_classification__isnull=True),
        status='active'
    ).count()

    context = {
        'heatmap_data': heatmap_data,
        'classifications': classifications,
        'high_risk_count': len(high_risk),
        'high_risk_no_runbook_count': len(high_risk_no_runbook),
        'high_risk_envs': high_risk[:50],
        'missing_data_count': missing_data,
        'total_active': Environment.objects.filter(status='active').count(),
    }

    return render(request, 'environments/risk_heatmap.html', context)


def service_primitives_inventory(request):
    """
    Service primitives inventory showing aggregation by service_primitive and service_class.
    Useful for capacity planning per service type.
    """
    from collections import defaultdict

    # Get filter parameters
    primitive_filter = request.GET.get('primitive')

    # Base queryset
    queryset = Environment.objects.all()

    # Apply filter
    if primitive_filter and primitive_filter != 'all':
        queryset = queryset.filter(service_primitive=primitive_filter)

    # Aggregate by service_primitive + service_class
    service_stats = defaultdict(lambda: {
        'service_primitive': '',
        'service_class': '',
        'total_envs': 0,
        'prod_count': 0,
        'staging_count': 0,
        'tier1_count': 0,
        'tier2_count': 0,
        'regions': defaultdict(int),
        'clouds': defaultdict(int),
    })

    for env in queryset.select_related():
        primitive = env.service_primitive or 'Unknown'
        service_class = env.service_class or 'Unknown'
        key = (primitive, service_class)
        stats = service_stats[key]
        stats['service_primitive'] = primitive
        stats['service_class'] = service_class
        stats['total_envs'] += 1

        if env.env_type == 'prod':
            stats['prod_count'] += 1
        elif env.env_type == 'staging':
            stats['staging_count'] += 1

        if env.criticality_tier == 1:
            stats['tier1_count'] += 1
        elif env.criticality_tier == 2:
            stats['tier2_count'] += 1

        if env.region:
            stats['regions'][env.region] += 1
        if env.cloud:
            stats['clouds'][env.cloud] += 1

    # Convert to sorted list
    service_list = []
    for key, stats in sorted(service_stats.items(), key=lambda x: x[1]['total_envs'], reverse=True):
        service_list.append({
            'service_primitive': stats['service_primitive'],
            'service_class': stats['service_class'],
            'total_envs': stats['total_envs'],
            'prod_count': stats['prod_count'],
            'staging_count': stats['staging_count'],
            'tier1_count': stats['tier1_count'],
            'tier2_count': stats['tier2_count'],
            'regions': dict(stats['regions']),
            'clouds': dict(stats['clouds']),
        })

    # Create table and configure for sorting/pagination
    table = ServicePrimitiveTable(service_list)
    RequestConfig(request, paginate={'per_page': 50}).configure(table)

    # Get filter choices
    primitives = [{'value': 'all', 'name': 'All Primitives'}]
    primitives.extend([
        {'value': p, 'name': p}
        for p in Environment.objects.values_list('service_primitive', flat=True).distinct().order_by('service_primitive')
        if p
    ])

    context = {
        'table': table,
        'services': service_list,
        'total_services': len(service_list),
        'total_envs': sum(s['total_envs'] for s in service_list),
        'primitives': primitives,
        'active_primitive': primitive_filter or 'all',
    }

    return render(request, 'environments/service_primitives.html', context)


# ---------------------------------------------------------------------------
# Service-category views (#DBaaS / CK8s aaS / Jenkins aaS / Builders)
# ---------------------------------------------------------------------------
from collections import Counter, defaultdict as _dd  # noqa: E402

# Builders span several naming conventions plus the builder_workloads flag.
BUILDER_Q = (
    Q(name__icontains='builder') | Q(name__icontains='autopkg')
    | Q(name__icontains='github-runner') | Q(name__icontains='lp-builder')
    | Q(name__icontains='job-farm-runner') | Q(name__icontains='superdistro')
    | Q(builder_workloads=True)
)


def _builder_subtype(name: str) -> str:
    n = (name or '').lower()
    if 'autopkg' in n:
        return 'autopkgtest'
    if 'superdistro' in n:
        return 'superdistro'
    if 'github-runner' in n or 'gh-runner' in n:
        return 'github-runner'
    if 'launchpad' in n or 'lp-builder' in n or 'job-farm-runner' in n:
        return 'launchpad-builder'
    return 'other-builder'


def service_base_queryset(slug):
    """Base queryset for a service slug — shared by the service pages and the
    env-list ?service= deep link so 'DBaaS for comsys' etc. filter identically."""
    if slug == 'dbaas':
        return Environment.objects.filter(service_class='database')
    if slug == 'ck8s':
        return Environment.objects.filter(k8s_distribution='ck8s')
    if slug == 'jenkins':
        return Environment.objects.filter(k8s_distribution='ck8s-jenkins-aas')
    if slug == 'vmaas':
        return Environment.objects.filter(service_class='machine_model')
    if slug == 'builders':
        return (Environment.objects.filter(BUILDER_Q)
                .exclude(name__startswith='objstor')
                .exclude(service_class='database'))
    return None


# Services shown on the high-level overview, in display order.
SERVICES = [
    ('dbaas', 'DBaaS'),
    ('ck8s', 'CK8s aaS'),
    ('vmaas', 'VMaaS / DevOps'),
    ('jenkins', 'Jenkins aaS'),
    ('builders', 'Builders'),
]


# ManSol (BootStack-managed) K8s clusters that CMDB actually tracks. The live
# clusters are named ``<env>-bootstack-managed-k8s-<X>`` on prodstack6; the infra
# repo references them under the rewritten name ``k8s-<env>-<X>``. Only clusters
# with at least one CMDB-modelled member model can be rendered here -- the other
# bootstack clusters (k8s-prod-cs-data-lakes, k8s-prod-cs-monitoring,
# k8s-prod-is-microk8s, k8s-prod-lma, k8s-{prod,stg}-mansol,
# k8s-{prod,stg}-se-devops) have no member models in CMDB and so cannot appear.
MANSOL_CLUSTERS = frozenset({
    'k8s-prod-general', 'k8s-prod-cs-general', 'k8s-prod-cs-trino',
    'k8s-prod-is', 'k8s-prod-snapstore',
    'k8s-stg-general', 'k8s-stg-cs-general', 'k8s-stg-is', 'k8s-stg-snapstore',
})


def k8s_clusters(request):
    """K8s clusters and the juju models inside them.

    Clusters are ``kubernetes_cluster`` envs (with an inline ``k8s_models``
    list); ``container_model`` envs link to their cluster via the ``cluster``
    field (the ManSol 'one file per model' pattern). The juju controller is the
    dependency link.

    Restricted to the ManSol (BootStack-managed) clusters listed in
    :data:`MANSOL_CLUSTERS`; all other K8s clusters are intentionally hidden.
    """
    from collections import defaultdict

    cloud_filter = request.GET.get('cloud')
    base = Environment.objects.filter(service_class='kubernetes_cluster')

    # container_model envs grouped by the cluster they point at.
    cmods = defaultdict(list)
    for e in (Environment.objects.filter(service_class='container_model')
              .exclude(cluster__isnull=True).exclude(cluster='').order_by('name')):
        cmods[e.cluster].append(e)
    cluster_names = set(base.values_list('name', flat=True))

    rows = []
    # IS-managed clusters (kubernetes_cluster envs with inline k8s_models).
    for c in base.order_by('cloud', 'name'):
        if c.name not in MANSOL_CLUSTERS:
            continue
        inline = [m.get('name') if isinstance(m, dict) else str(m)
                  for m in (c.k8s_models or [])]
        containers = cmods.get(c.name, [])
        rows.append({
            'kind': 'managed', 'name': c.name, 'cloud': c.cloud, 'env': c,
            'mansol': True,
            'controller': c.juju_controller, 'managed_by': c.owner,
            'cp': (f"{c.control_plane_size}×{c.control_plane_units}"
                   if c.control_plane_size else ''),
            'workers': ', '.join(f"{g.get('size')}×{g.get('units')}"
                                 for g in (c.worker_groups or [])),
            'inline': inline, 'containers': containers,
            'model_count': len(inline) + len(containers),
        })
    # External / ManSol clusters: referenced by container_models but not an env
    # here (managed by another team). Synthesize from their member models.
    for cname, members in cmods.items():
        if cname in cluster_names:
            continue
        if cname not in MANSOL_CLUSTERS:
            continue
        m0 = members[0]
        rows.append({
            'kind': 'external', 'name': cname, 'cloud': m0.cloud, 'env': None,
            'mansol': True,
            'controller': m0.juju_controller, 'managed_by': m0.owner or '—',
            'cp': '', 'workers': '', 'inline': [], 'containers': members,
            'model_count': len(members),
        })

    # Dropdown reflects only the clouds the ManSol clusters live on (ps6).
    mansol_clouds = sorted({r['cloud'] for r in rows if r['cloud']})
    if cloud_filter and cloud_filter != 'all':
        rows = [r for r in rows if r['cloud'] == cloud_filter]

    # Server-side column sorting (persists in the URL, works without JS, and
    # sorts the whole result set rather than just the current DOM).
    columns = [
        {'key': 'name', 'label': 'Cluster'},
        {'key': 'cloud', 'label': 'Cloud'},
        {'key': 'controller', 'label': 'Juju controller'},
        {'key': 'managed_by', 'label': 'Managed by'},
        {'key': 'cp', 'label': 'Control plane'},
        {'key': 'workers', 'label': 'Workers'},
        {'key': 'models', 'label': 'Models', 'num': True},
        {'key': 'juju_models', 'label': 'Juju models'},
    ]
    sort_keys = {
        'name': lambda r: (r['name'] or '').lower(),
        'cloud': lambda r: (r['cloud'] or '', (r['name'] or '').lower()),
        'controller': lambda r: (r['controller'] or '').lower(),
        'managed_by': lambda r: (r['managed_by'] or '').lower(),
        'cp': lambda r: (r['cp'] or '').lower(),
        'workers': lambda r: (r['workers'] or '').lower(),
        'models': lambda r: r['model_count'],
        # The "Juju models" column lists names; sort by them alphabetically
        # (distinct from the numeric "Models" count column).
        'juju_models': lambda r: ' '.join([str(m).lower() for m in r['inline']]
                                          + [e.name.lower() for e in r['containers']]),
    }
    sort = request.GET.get('sort')
    if sort not in sort_keys:
        sort = 'cloud'
    direction = 'desc' if request.GET.get('dir') == 'desc' else 'asc'
    rows.sort(key=sort_keys[sort], reverse=(direction == 'desc'))

    clouds = [{'value': 'all', 'name': 'All Clouds'}]
    clouds += [{'value': cl, 'name': cl} for cl in mansol_clouds]

    return render(request, 'environments/k8s_clusters.html', {
        'rows': rows, 'clouds': clouds, 'active_cloud': cloud_filter or 'all',
        'total': len(rows), 'columns': columns, 'sort': sort, 'dir': direction,
    })


def k8s_cluster_detail(request, name):
    """Detail for one K8s cluster: the juju models inside it plus cluster metadata.

    Mirrors the two cluster kinds in :func:`k8s_clusters`: an IS-managed
    ``kubernetes_cluster`` env (carrying an inline ``k8s_models`` list, control
    plane and worker groups), or an external/ManSol cluster synthesised from the
    ``container_model`` envs whose ``cluster`` field points at it.
    """
    from django.http import Http404
    if name not in MANSOL_CLUSTERS:
        raise Http404(f"{name!r} is not a ManSol K8s cluster")

    cluster_env = (Environment.objects
                   .filter(service_class='kubernetes_cluster', name=name).first())
    containers = list(Environment.objects
                      .filter(service_class='container_model', cluster=name)
                      .order_by('name'))
    if cluster_env is None and not containers:
        raise Http404(f"No K8s cluster named {name!r}")

    # Inline k8s_models declared on a managed cluster env (dicts or bare names).
    inline = []
    for m in (getattr(cluster_env, 'k8s_models', None) or []):
        if isinstance(m, dict):
            inline.append({'name': m.get('name') or '', 'meta': m})
        else:
            inline.append({'name': str(m), 'meta': {}})

    kind = 'managed' if cluster_env else 'external'
    src = cluster_env or containers[0]
    workers = [f"{g.get('size')}×{g.get('units')}"
               for g in (getattr(cluster_env, 'worker_groups', None) or [])]

    # "Whatever else we got": the clouds / owners / teams the members span.
    member_clouds = {c.cloud for c in containers if c.cloud}
    if cluster_env and cluster_env.cloud:
        member_clouds.add(cluster_env.cloud)
    owners = sorted({c.owner for c in containers if c.owner})
    teams = sorted({c.team for c in containers if c.team})

    meta = {
        'kind': kind,
        'cloud': src.cloud,
        'controller': src.juju_controller,
        'managed_by': (cluster_env.owner if cluster_env else (containers[0].owner or '—')),
        'stage': (cluster_env.juju_controller_stage if cluster_env else None) or src.env_type,
        'juju_series': src.juju_series,
        'bastion': getattr(src, 'bastion_server', None),
        'risk_group': src.risk_group,
        'k8s_distribution': getattr(cluster_env, 'k8s_distribution', None),
        'control_plane': (f"{cluster_env.control_plane_size}×{cluster_env.control_plane_units}"
                          if cluster_env and cluster_env.control_plane_size else ''),
        'workers': ', '.join(workers),
        'description': src.description,
        'status': getattr(cluster_env, 'status', None),
        'quotas': (cluster_env.quotas if cluster_env else None) or {},
    }

    # CIA-assessment stakeholders, aggregated from the cluster's member models
    # (the ``cia_assessment.asset`` block: owner / risk_owner / custodian). The
    # source YAML also has a ``delegate`` but the parser does not persist it.
    from collections import Counter
    cia_envs = containers + ([cluster_env] if cluster_env else [])
    risk_owners = Counter(e.cia_risk_owner for e in cia_envs if e.cia_risk_owner)
    cia_owners = Counter(e.cia_owner for e in cia_envs if e.cia_owner)
    custodians = Counter(e.cia_custodian for e in cia_envs if e.cia_custodian)
    cia = {
        'risk_owners': [{'value': v, 'count': n} for v, n in risk_owners.most_common()],
        'owners': [{'value': v, 'count': n} for v, n in cia_owners.most_common()],
        'custodians': [{'value': v, 'count': n} for v, n in custodians.most_common()],
        'models_with_cia': sum(1 for e in cia_envs
                               if e.cia_owner or e.cia_risk_owner or e.cia_custodian),
        'models_total': len(cia_envs),
        'has_any': bool(risk_owners or cia_owners or custodians),
    }

    return render(request, 'environments/k8s_cluster_detail.html', {
        'name': name, 'meta': meta, 'cluster_env': cluster_env, 'mansol': True,
        'containers': containers, 'inline': inline,
        'model_count': len(inline) + len(containers),
        'clouds': sorted(member_clouds), 'owners': owners, 'teams': teams,
        'cia': cia,
    })


def juju_controllers(request):
    """Juju controllers + JAAS/JIMM across all clouds with quota and placement."""
    from cmdb.redis_client import placement_map
    import re as _re

    # Bootstrap controllers (juju_controller), the JAAS controller mesh, and the
    # JIMM deployments (clusters running the juju-jimm-k8s charm).
    base = Environment.objects.filter(
        Q(service_class__icontains='juju_controller')
        | Q(service_class__icontains='jaas_controller')
        | Q(charm_versions__has_key='juju-jimm-k8s')
    )
    cloud_filter = request.GET.get('cloud')
    qs = base
    if cloud_filter and cloud_filter != 'all':
        qs = qs.filter(cloud=cloud_filter)
    qs = qs.order_by('cloud', 'name')

    names = list(qs.values_list('name', flat=True))
    pmap = placement_map(names)

    def short_node(host, cloud):
        s = (host or '').split('.')[0]
        if cloud and s.startswith(cloud + '-'):
            return s[len(cloud) + 1:]
        return _re.sub(r'^ps\d+-', '', s)

    rows = []
    totals = {'cores': 0, 'ram_gb': 0.0, 'disk': 0, 'vm_count': 0}
    for env in qs:
        q = env.quotas or {}
        p = pmap.get(env.name) or {}
        hosts = p.get('hosts') or []
        cloud = p.get('cloud') or env.cloud or ''
        nodes = [short_node(h, cloud) for h in hosts][:3]
        cores = q.get('cores') or 0
        ram_mb = q.get('ram') or 0
        disk = q.get('gigabytes') or 0
        vm_count = p.get('vm_count') or 0
        ram_gb = round(ram_mb / 1024, 1)
        totals['cores'] += cores
        totals['ram_gb'] += ram_gb
        totals['disk'] += disk
        totals['vm_count'] += vm_count
        rows.append({
            'name': env.name, 'cloud': env.cloud, 'region': env.region,
            'consumed_by': env.consumed_by, 'is_jimm': 'juju-jimm-k8s' in (env.charm_versions or {}),
            'cores': q.get('cores'), 'ram_gb': ram_gb if ram_mb else None, 'disk': q.get('gigabytes'),
            'nodes': ', '.join(nodes), 'extra_nodes': max(0, len(hosts) - 3),
            'vm_count': vm_count,
        })
    totals['ram_gb'] = round(totals['ram_gb'], 1)

    clouds = [{'value': 'all', 'name': 'All Clouds'}]
    clouds += [{'value': c, 'name': c} for c in base.exclude(cloud__isnull=True)
               .exclude(cloud='').values_list('cloud', flat=True).distinct().order_by('cloud')]

    return render(request, 'environments/juju_controllers.html', {
        'rows': rows, 'clouds': clouds, 'active_cloud': cloud_filter or 'all',
        'total': len(rows), 'totals': totals,
    })


def services_overview(request):
    """High-level overview: env count + *consuming* teams per service.

    The consumer is the requester (for aaS, team/owner is the provider 'is'),
    falling back to the owning team — see Environment.consumed_by.
    """
    cards = []
    for slug, label in SERVICES:
        qs = service_base_queryset(slug)
        consumers = []
        if qs is not None:
            seen = set()
            for requester, team in qs.values_list('requester', 'team'):
                c = (requester or '').strip() or (team or '').strip()
                if c:
                    seen.add(c)
            consumers = sorted(seen)
        cards.append({
            'slug': slug, 'label': label,
            'count': qs.count() if qs is not None else 0,
            'consumers': consumers, 'consumer_count': len(consumers),
        })
    return render(request, 'environments/services_overview.html', {'cards': cards})


def _service_view(request, *, title, slug, base_qs, group_field, group_label):
    """Shared aggregation for a service category: team usage, group(size)
    breakdown, and total resource usage. ``group_field`` is an Environment
    attribute name, or the sentinel '__builder_subtype__'."""
    region_filter = request.GET.get('region')
    cloud_filter = request.GET.get('cloud')

    qs = base_qs
    if region_filter and region_filter != 'all':
        qs = qs.filter(region=region_filter)
    if cloud_filter and cloud_filter != 'all':
        qs = qs.filter(cloud=cloud_filter)

    teams = _dd(lambda: {'count': 0, 'cores': 0, 'ram': 0, 'instances': 0, 'gigabytes': 0})
    groups = Counter()
    totals = {'count': 0, 'cores': 0, 'ram': 0, 'instances': 0, 'gigabytes': 0}
    env_rows = []

    for env in qs.order_by('name'):
        q = env.quotas or {}
        # Group by the consuming team (requester for aaS; else owning team),
        # not the provider 'is'.
        t = teams[env.consumed_by or 'Unknown']
        for k in ('cores', 'ram', 'instances', 'gigabytes'):
            v = q.get(k, 0) or 0
            t[k] += v
            totals[k] += v
        t['count'] += 1
        totals['count'] += 1
        if group_field == '__builder_subtype__':
            grp = _builder_subtype(env.name)
        else:
            grp = getattr(env, group_field, None) or 'unspecified'
        groups[grp] += 1
        env_rows.append({
            'name': env.name, 'team': env.team, 'owner': env.owner,
            'cloud': env.cloud, 'group': grp, 'status': env.status,
            'cores': q.get('cores', 0) or 0,
            'ram_gb': round((q.get('ram', 0) or 0) / 1024, 1),
            'instances': q.get('instances', 0) or 0,
        })

    teams_list = sorted(
        [
            {'team': k, 'count': v['count'], 'cores': v['cores'],
             'ram_gb': round(v['ram'] / 1024, 1), 'instances': v['instances'],
             'storage_gb': v['gigabytes']}
            for k, v in teams.items()
        ],
        key=lambda x: x['cores'], reverse=True,
    )
    groups_list = sorted(
        [{'name': k, 'count': c} for k, c in groups.items()],
        key=lambda x: x['count'], reverse=True,
    )

    regions = [{'value': 'all', 'name': 'All Regions'}]
    regions += [{'value': r, 'name': r.upper()}
                for r in base_qs.exclude(region__isnull=True).exclude(region='')
                .values_list('region', flat=True).distinct().order_by('region')]
    clouds = [{'value': 'all', 'name': 'All Clouds'}]
    clouds += [{'value': c, 'name': c}
               for c in base_qs.exclude(cloud__isnull=True).exclude(cloud='')
               .values_list('cloud', flat=True).distinct().order_by('cloud')]

    context = {
        'title': title,
        'slug': slug,
        'teams': teams_list,
        'groups': groups_list,
        'group_label': group_label,
        'environments': env_rows,
        'totals': {**totals, 'ram_gb': round(totals['ram'] / 1024, 1),
                   'storage_gb': totals['gigabytes'], 'teams': len(teams_list)},
        'regions': regions,
        'clouds': clouds,
        'active_region': region_filter or 'all',
        'active_cloud': cloud_filter or 'all',
    }
    return render(request, 'environments/service_view.html', context)


def dbaas_view(request):
    return _service_view(
        request, title='DBaaS', slug='dbaas',
        base_qs=service_base_queryset('dbaas'),
        group_field='database_size', group_label='Database size',
    )


def ck8s_aas_view(request):
    return _service_view(
        request, title='CK8s aaS', slug='ck8s',
        base_qs=service_base_queryset('ck8s'),
        group_field='control_plane_size', group_label='Control-plane size',
    )


def jenkins_aas_view(request):
    return _service_view(
        request, title='Jenkins aaS', slug='jenkins',
        base_qs=service_base_queryset('jenkins'),
        group_field='control_plane_size', group_label='Control-plane size',
    )


def builders_view(request):
    return _service_view(
        request, title='Builders', slug='builders',
        base_qs=service_base_queryset('builders'),
        group_field='__builder_subtype__', group_label='Builder type',
    )


# ---------------------------------------------------------------------------
# GitOps mapping dashboard
# ---------------------------------------------------------------------------
def gitops_teams(request):
    """GitOps teams sub-tab: per-team count and percentage of GitOps-managed
    environments, as two bar charts."""
    rows = list(
        Environment.objects.exclude(team__isnull=True).exclude(team='')
        .values('team')
        .annotate(total=Count('id'),
                  gitops=Count('id', filter=Q(gitops_managed=True)))
    )
    for r in rows:
        r['pct'] = round(r['gitops'] / r['total'] * 100, 1) if r['total'] else 0.0

    by_count = sorted([r for r in rows if r['gitops']],
                      key=lambda r: r['gitops'], reverse=True)[:20]
    by_pct = sorted([r for r in rows if r['total']],
                    key=lambda r: r['pct'], reverse=True)[:20]
    max_count = max([r['gitops'] for r in by_count], default=1) or 1

    return render(request, 'environments/gitops_teams.html', {
        'by_count': by_count,
        'by_pct': by_pct,
        'max_count': max_count,
    })


def gitops_overview(request):
    """
    Map infrastructure-services environments to their Terraform models
    (is-terraform-models et al.) and the reusable is-terraform-modules they
    consume. The GitOps signal comes from Environment.gitops_model_management
    (flattened by `refresh_gitops`). An optional ?repo= narrows to one repo.
    """
    repo_filter = request.GET.get('repo')

    managed_qs = Environment.objects.filter(gitops_managed=True)
    total = Environment.objects.count()
    managed_count = managed_qs.count()

    # Per-repository breakdown (count / enabled / suspended).
    repos = {}
    for r in managed_qs.values('gitops_repo').annotate(
        count=Count('id'),
        enabled=Count('id', filter=Q(gitops_enabled=True)),
        suspended=Count('id', filter=Q(gitops_suspended=True)),
    ).order_by('-count'):
        repos[r['gitops_repo'] or 'unknown'] = r
    repo_list = list(repos.values())

    # Apply the repo filter to the env/module tables (but not the summary cards).
    env_qs = managed_qs
    if repo_filter:
        env_qs = env_qs.filter(gitops_repo=repo_filter)

    # Module-usage tally across the (optionally filtered) managed envs.
    module_tally = Counter()
    managed_envs = []
    for env in env_qs.order_by('gitops_repo', 'name').values(
        'name', 'gitops_repo', 'gitops_path', 'gitops_enabled',
        'gitops_suspended', 'gitops_modules', 'env_type',
    ):
        mods = env.get('gitops_modules') or []
        module_tally.update(mods)
        managed_envs.append(env)

    module_list = [
        {'module': m, 'count': c} for m, c in module_tally.most_common()
    ]

    context = {
        'total': total,
        'managed_count': managed_count,
        'unmanaged_count': total - managed_count,
        'repo_count': len(repo_list),
        'enabled_count': managed_qs.filter(gitops_enabled=True).count(),
        'suspended_count': managed_qs.filter(gitops_suspended=True).count(),
        'repo_list': repo_list,
        'managed_envs': managed_envs,
        'module_list': module_list,
        'modules_resolved_count': sum(1 for e in managed_envs if e.get('gitops_modules')),
        'repo_filter': repo_filter,
    }
    return render(request, 'environments/gitops_overview.html', context)
